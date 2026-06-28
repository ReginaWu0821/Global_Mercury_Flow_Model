from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell


TARGET_PREFIXES = frozenset({"1", "2", "3", "4", "5"})


@dataclass
class FreezeStats:
    workbook: str
    scanned_cells: int = 0
    replaced_formula_cells: int = 0
    copied_constant_cells: int = 0
    missing_formula_cache: int = 0
    remaining_formula_cells: int = 0
    sample_missing_formula_cells: list[str] | None = None

    def __post_init__(self) -> None:
        if self.sample_missing_formula_cells is None:
            self.sample_missing_formula_cells = []


def _iter_target_workbooks(root: Path,
                           prefixes: frozenset[str] = TARGET_PREFIXES):
    for wb_path in sorted(root.glob("*.xlsx")):
        if wb_path.name.startswith("~$"):
            continue
        if wb_path.name[:1] in prefixes:
            yield wb_path


def _count_formulas(wb) -> int:
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.data_type == "f":
                    total += 1
    return total


def freeze_workbook_to_values(wb_path: Path,
                              *,
                              fail_on_missing_cache: bool = True,
                              max_missing_samples: int = 20) -> FreezeStats:
    """Replace all visible cell values with their cached display values.

    The workbook is updated in place. Formula cells are replaced by their
    cached results from ``data_only=True``. Non-formula cells are also written
    back from the cached view so the whole workbook is value-backed.
    """

    wb_path = Path(wb_path)
    wb_formula = openpyxl.load_workbook(wb_path, data_only=False)
    wb_values = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    stats = FreezeStats(workbook=wb_path.name)

    try:
        for ws_formula in wb_formula.worksheets:
            if ws_formula.title not in wb_values.sheetnames:
                continue
            ws_values = wb_values[ws_formula.title]

            # 用并集范围避免漏到任一侧的尾部行/列；和原 range(1..max) 等价。
            max_row = max(ws_formula.max_row, ws_values.max_row)
            max_col = max(ws_formula.max_column, ws_values.max_column)

            # 用 iter_rows 替代 range × ws.cell() 的随机访问：
            # - 写入侧（普通模式）仍能拿到与 ws.cell(r,c) 相同的 Cell 对象，
            #   写 cell.value 直接落到 ws._cells；
            # - 读取侧是 read_only 工作簿，避免随机访问触发的反复重扫；
            #   两侧给同样的 (min/max_row, min/max_col)，每行长度严格相等，
            #   因此 zip 不会错位，逐 (r,c) 的语义与原实现一致。
            rows_formula = ws_formula.iter_rows(
                min_row=1, max_row=max_row,
                min_col=1, max_col=max_col,
            )
            rows_values = ws_values.iter_rows(
                min_row=1, max_row=max_row,
                min_col=1, max_col=max_col,
            )

            for row_formula, row_values in zip(rows_formula, rows_values):
                for cell_formula, cell_values in zip(row_formula, row_values):
                    if isinstance(cell_formula, MergedCell):
                        continue

                    stats.scanned_cells += 1
                    cell_value = cell_values.value

                    if cell_formula.data_type == "f":
                        if cell_value is None:
                            stats.missing_formula_cache += 1
                            if len(stats.sample_missing_formula_cells) < max_missing_samples:
                                stats.sample_missing_formula_cells.append(
                                    f"{ws_formula.title}!{cell_formula.coordinate}"
                                )
                            continue
                        cell_formula.value = cell_value
                        stats.replaced_formula_cells += 1
                    else:
                        if cell_formula.value != cell_value:
                            stats.copied_constant_cells += 1
                        cell_formula.value = cell_value

        if fail_on_missing_cache and stats.missing_formula_cache:
            samples = ", ".join(stats.sample_missing_formula_cells or [])
            raise RuntimeError(
                f"{wb_path.name}: {stats.missing_formula_cache} formula cell(s) "
                f"missing cached values; samples: {samples}"
            )

        wb_formula.save(wb_path)
    finally:
        wb_values.close()
        wb_formula.close()

    wb_check = openpyxl.load_workbook(wb_path, data_only=False, read_only=True)
    try:
        stats.remaining_formula_cells = _count_formulas(wb_check)
    finally:
        wb_check.close()
    return stats


def freeze_directory_workbooks(root: Path,
                               *,
                               prefixes: frozenset[str] = TARGET_PREFIXES,
                               fail_on_missing_cache: bool = True) -> list[FreezeStats]:
    root = Path(root)
    results: list[FreezeStats] = []
    for wb_path in _iter_target_workbooks(root, prefixes=prefixes):
        results.append(
            freeze_workbook_to_values(
                wb_path,
                fail_on_missing_cache=fail_on_missing_cache,
            )
        )
    return results

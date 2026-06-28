from __future__ import annotations

import os
import re
from pathlib import Path

import openpyxl

from _wb_resolver import INPUT_DIR_ENV, resolve_workbook_path

_DIRECT_REF_RE = re.compile(
    r"^=\s*(?:\[[^\]]+\])?(?:'[^']+'|[A-Za-z0-9_ .-]+)?!?\$?[A-Z]{1,3}\$?\d+\s*$"
)
_FUNC_RE = re.compile(r"[A-Z][A-Z0-9_]*\(")


def _canonical_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.replace("-2022", "")).strip().lower()


def _resolve_input_source(repo_root: Path, workbook_name: str) -> Path:
    input_dir = os.environ.get(INPUT_DIR_ENV, "input_v2")
    src_root = Path(input_dir).expanduser()
    if not src_root.is_absolute():
        src_root = (repo_root / src_root).resolve()

    exact = src_root / workbook_name
    if exact.exists():
        return exact

    target = _canonical_name(workbook_name)
    for cand in sorted(src_root.glob("*.xlsx")):
        if _canonical_name(cand.name) == target:
            return cand

    raise FileNotFoundError(
        f"Cannot resolve source workbook '{workbook_name}' under {src_root}"
    )


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _is_direct_formula(formula: str) -> bool:
    return bool(_DIRECT_REF_RE.match(formula))


def _is_complex_formula(formula: str) -> bool:
    if _is_direct_formula(formula):
        return False
    core = formula[1:]
    if _FUNC_RE.search(core):
        return True
    if re.search(r"[\+\-\*/\^&]", core):
        return True
    if ":" in core or "," in core:
        return True
    return False


def sync_complex_formulas(repo_root: Path, workbook_name: str, sheet_name: str) -> None:
    src_path = _resolve_input_source(repo_root, workbook_name)
    dst_path = resolve_workbook_path(repo_root, workbook_name)

    src_wb = openpyxl.load_workbook(src_path, data_only=False, read_only=True)
    if sheet_name not in src_wb.sheetnames:
        src_wb.close()
        raise KeyError(f"Sheet '{sheet_name}' not found in source workbook {src_path}")

    dst_wb = openpyxl.load_workbook(dst_path)
    if sheet_name not in dst_wb.sheetnames:
        src_wb.close()
        dst_wb.close()
        raise KeyError(f"Sheet '{sheet_name}' not found in target workbook {dst_path}")

    src_ws = src_wb[sheet_name]
    dst_ws = dst_wb[sheet_name]
    changed = False

    for row in src_ws.iter_rows():
        for cell in row:
            value = cell.value
            if _is_formula(value) and _is_complex_formula(value):
                src_formula = str(value)
                dst_cell = dst_ws.cell(cell.row, cell.column)
                if str(dst_cell.value) != src_formula:
                    dst_cell.value = src_formula
                    changed = True

    if changed:
        dst_wb.save(dst_path)
    src_wb.close()
    dst_wb.close()

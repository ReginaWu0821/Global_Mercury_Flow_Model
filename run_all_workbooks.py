"""总控：按工作簿顺序为 ``Code/`` 下每个模块目录调用共享的 sheet runner。

I/O 行为
========
* ``--input-dir/-i``：指定输入工作簿所在目录（默认 ``input``）。
  原始输入文件全程**只读**，不会被修改。
* ``--output-dir/-o``：指定输出目录（默认 ``output``）。运行时会先把输入工作簿
  复制一份到该目录，所有子脚本里的 ``wb.save(...)`` 都会落到这份副本上。
  ``--keep-output`` 关闭运行前的清空行为。

具体路径解析与"输入→输出复制"由 ``Code/_wb_resolver.py`` 通过环境变量
``GMF_INPUT_DIR`` / ``GMF_OUTPUT_DIR`` 完成；各子模块脚本无需改动。
"""

from __future__ import annotations

import argparse
import functools
import os
import re
import shutil
import subprocess
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent / "Code"))

from _formula_value_freezer import freeze_directory_workbooks


def _format_duration(seconds: float) -> str:
    """把秒数格式化为 ``HH:MM:SS.mmm`` 与原始秒数并列展示。"""
    total_ms = int(round(seconds * 1000))
    h, rem = divmod(total_ms, 3600 * 1000)
    m, rem = divmod(rem, 60 * 1000)
    s, ms = divmod(rem, 1000)
    return f"{h:02d}:{m:02d}:{s:02d}.{ms:03d} ({seconds:.3f}s)"


def timed(label: str | None = None):
    """计时装饰器：在被装饰函数返回（含异常）后打印耗时。"""

    def decorator(func):
        name = label or func.__name__

        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            start = time.perf_counter()
            try:
                return func(*args, **kwargs)
            finally:
                elapsed = time.perf_counter() - start
                print(f"\n[TIMING] {name} elapsed: {_format_duration(elapsed)}")

        return wrapper

    return decorator


BOOK_RE = re.compile(r"工作簿:\s*(?:Global_Hg_Flow|Code-2022)/(.+?\.xlsx)")
SHEET_RE = re.compile(r"工作表:\s*'([^']+)'")


def extract_workbook_from_folder(folder: Path) -> str | None:
    for script in sorted(folder.glob('*.py')):
        if script.name.startswith('_'):
            continue
        text = script.read_text(encoding='utf-8', errors='ignore')
        mb = BOOK_RE.search(text)
        ms = SHEET_RE.search(text)
        if mb and ms:
            return mb.group(1)
    return None


def workbook_sort_key(name: str) -> tuple[int, str]:
    norm = name.replace("-2022", "")
    norm = re.sub(r"\s+", " ", norm).strip()
    m = re.match(r"(\d+)(?:\.(\d+))?", norm)
    if not m:
        return (9999, norm)
    major = int(m.group(1))
    minor = int(m.group(2)) if m.group(2) is not None else -1
    return (major * 100 + (minor if minor >= 0 else 0), norm)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="按工作簿顺序运行 Code/ 下各模块"
    )
    parser.add_argument(
        "-i", "--input-dir",
        default="input",
        help="输入工作簿所在目录（相对路径基于仓库根目录解析），默认 input",
    )
    parser.add_argument(
        "-o", "--output-dir",
        default="output",
        help="输出目录，所有写回操作落到该目录下的工作簿副本，默认 output",
    )
    parser.add_argument(
        "--keep-output",
        action="store_true",
        help="运行前不清空输出目录（默认会清空，避免上一次的旧产物干扰）",
    )
    return parser.parse_args(argv)


def _inline_cached_values(wb, wb_cached, keep_sheets: list[str], wb_label: str) -> int:
    """把 ``keep_sheets`` 里所有公式格替换为 ``wb_cached`` 中对应的缓存计算值。

    背景
    ====
    后处理会删掉一部分 sheet，但保留下来的 sheet（如 ``Sum_*``）里的公式
    经常引用被删 sheet 的内容（如 ``=VLOOKUP($C4,DF_CW!$A$3:$F$14,...)``）。
    一旦被引用 sheet 被删，这些公式在 Excel 里重新打开就变成 ``#REF!`` /
    显示为 NA。解决办法：删除前，把保留 sheet 里的公式格换成它们的"渲染值"
    （缓存计算结果），让保留 sheet 在丢失依赖后仍然显示原本算出来的数。

    实现
    ====
    * 公式格的判定：``cell.data_type == 'f'``（含数组公式 anchor）。
    * 缓存值来自同名工作簿的 ``data_only=True`` 副本。``data_only`` 视图里
      公式格会直接呈现 Excel 上次保存时缓存的计算结果。
    * 缓存为 ``None`` 的格子：保留原公式并打印 WARN，避免误把它清空成
      空白（这恰是用户想避免的失败模式）。在当前项目里这种情况非常罕见
      （扫过整套 ``output_answer/`` 中保留 sheet 的所有公式格，缓存覆盖
      率 100%），但兜底逻辑仍保留以防 input 端有缺漏。

    返回
    ====
    实际固化（替换）的公式格数量。
    """

    inlined = 0
    no_cache = 0
    for sheet_name in keep_sheets:
        ws = wb[sheet_name]
        ws_cached = wb_cached[sheet_name] if sheet_name in wb_cached.sheetnames else None
        if ws_cached is None:
            print(f"[POSTPROCESS][WARN] {wb_label}: sheet '{sheet_name}' missing in data_only view; skipped inlining")
            continue

        for row_f, row_d in zip(ws.iter_rows(), ws_cached.iter_rows()):
            for cell, cached in zip(row_f, row_d):
                if cell.data_type != "f":
                    continue
                cached_val = cached.value
                if cached_val is None:
                    no_cache += 1
                    continue
                try:
                    cell.value = cached_val
                except (AttributeError, TypeError) as exc:
                    print(f"[POSTPROCESS][WARN] {wb_label}: cannot inline "
                          f"{sheet_name}!{cell.coordinate} ({exc})")
                    continue
                inlined += 1

    if no_cache:
        print(f"[POSTPROCESS][WARN] {wb_label}: {no_cache} formula cell(s) had no cached value; left as formula")
    return inlined


def _postprocess_outputs(output_dir: Path) -> None:
    """对输出目录里的工作簿做最终精简：

    * 文件名首字符为 ``1``-``4`` 的 xlsx：只保留 sheet 名以 ``sum``（不区分大小写）
      开头的工作表；同时检查这些被保留的 sheet 是否带 tab 颜色，若没有则打印
      警告（按约定它们应该都是标了颜色的"汇总"sheet）。
    * 文件名首字符为 ``5`` 的 xlsx：删掉名为 ``DF_Waste`` 的工作表。

    其它文件保持不变。

    由于保留 sheet 中的公式经常跨 sheet 引用即将被删的 sheet（典型例子：
    ``Sum_CW_CFPP!K4 = VLOOKUP($C4, DF_CW!$A$3:$F$14, ...)``），如果直接删掉
    ``DF_CW`` 之类的 sheet，再次打开工作簿这些公式就会变成 ``#REF!``。所以在
    删除前会先把保留 sheet 的公式格替换为缓存计算值（见 ``_inline_cached_values``）。
    """

    import openpyxl  # 局部导入，避免影响入口的轻量启动

    print("\n=== [POSTPROCESS] trimming output workbooks ===")
    for wb_path in sorted(output_dir.glob("*.xlsx")):
        first_char = wb_path.name[:1]
        if first_char in {"1", "2", "3", "4"}:
            mode = "keep_sum"
        elif first_char == "5":
            mode = "drop_df_waste"
        else:
            continue

        try:
            wb = openpyxl.load_workbook(wb_path)
            wb_cached = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        except Exception as exc:  # 读不开就跳过，不阻断整个流程
            print(f"[POSTPROCESS][WARN] cannot open {wb_path.name}: {exc}")
            continue

        original_sheets = list(wb.sheetnames)
        changed = False

        if mode == "keep_sum":
            keep = [s for s in original_sheets if s.lower().startswith("sum")]
            if not keep:
                print(f"[POSTPROCESS][WARN] {wb_path.name}: no sheet starts with 'sum'; left untouched")
                wb.close()
                wb_cached.close()
                continue

            # 颜色检查：有颜色的 sheet 期望就是这些 sum 开头的 sheet。
            for name in keep:
                tab = wb[name].sheet_properties.tabColor
                if tab is None or not getattr(tab, "rgb", None):
                    print(f"[POSTPROCESS][WARN] {wb_path.name}: kept sheet '{name}' has no tab color")

            # 删 sheet 之前先把保留 sheet 中跨 sheet 公式格固化为缓存值，
            # 否则被删 sheet 一走，留下来的公式就全是 #REF!/NA 了。
            inlined = _inline_cached_values(wb, wb_cached, keep, wb_path.name)
            if inlined:
                changed = True

            for name in original_sheets:
                if name not in keep:
                    del wb[name]
                    changed = True
            print(f"[POSTPROCESS] {wb_path.name}: inlined {inlined} formula cell(s) "
                  f"across {len(keep)} kept sheet(s)")

        elif mode == "drop_df_waste":
            target = "DF_Waste"
            if target in wb.sheetnames:
                # 同样的考虑：保留 sheet 里可能引用 DF_Waste，先固化它们的
                # 缓存值，再删 DF_Waste。
                keep = [s for s in original_sheets if s != target]
                inlined = _inline_cached_values(wb, wb_cached, keep, wb_path.name)
                del wb[target]
                changed = True
                print(f"[POSTPROCESS] {wb_path.name}: inlined {inlined} formula cell(s) "
                      f"across {len(keep)} kept sheet(s)")
            else:
                print(f"[POSTPROCESS][INFO] {wb_path.name}: no '{target}' sheet to drop")

        # 缓存视图先关掉，再写回（避免 read_only 句柄持有同一份 zip）。
        wb_cached.close()

        if changed:
            wb.save(wb_path)
            print(f"[POSTPROCESS] {wb_path.name}: "
                  f"{len(original_sheets)} -> {len(wb.sheetnames)} sheets")
        wb.close()


def _freeze_output_formulas(output_dir: Path) -> None:
    print("\n=== [POSTPROCESS] freezing formulas to cached values ===")
    results = freeze_directory_workbooks(output_dir, fail_on_missing_cache=True)

    total_scanned = 0
    total_formula = 0
    total_constant = 0
    total_missing = 0
    total_remaining = 0
    for stats in results:
        total_scanned += stats.scanned_cells
        total_formula += stats.replaced_formula_cells
        total_constant += stats.copied_constant_cells
        total_missing += stats.missing_formula_cache
        total_remaining += stats.remaining_formula_cells
        print(
            f"[FREEZE] {stats.workbook}: scanned={stats.scanned_cells}, "
            f"formula_to_value={stats.replaced_formula_cells}, "
            f"constant_rewritten={stats.copied_constant_cells}, "
            f"missing_cache={stats.missing_formula_cache}, "
            f"remaining_formulas={stats.remaining_formula_cells}"
        )

    print(
        f"[FREEZE][DONE] workbooks={len(results)}, scanned={total_scanned}, "
        f"formula_to_value={total_formula}, constant_rewritten={total_constant}, "
        f"missing_cache={total_missing}, remaining_formulas={total_remaining}"
    )


def _resolve_under_repo(repo: Path, p: str) -> Path:
    path = Path(p).expanduser()
    if not path.is_absolute():
        path = (repo / path).resolve()
    return path


def _prepare_output_dir(output_dir: Path, repo: Path, keep: bool) -> int:
    """Create / clean the output directory. Returns 0 on success, non-zero on error."""

    repo_resolved = repo.resolve()
    out_resolved = output_dir.resolve() if output_dir.exists() else output_dir

    # 安全检查：不允许把 output_dir 设成仓库根，或仓库根的父目录
    # （清理这种目录可能误删源代码）。
    if out_resolved == repo_resolved:
        print(f"[ERROR] refuse to use repo root as output dir: {output_dir}",
              file=sys.stderr)
        return 2
    try:
        if repo_resolved.is_relative_to(out_resolved):
            print(f"[ERROR] refuse to use an ancestor of repo as output dir: {output_dir}",
                  file=sys.stderr)
            return 2
    except (AttributeError, ValueError):
        pass

    if output_dir.exists() and not output_dir.is_dir():
        print(f"[ERROR] output path exists and is not a directory: {output_dir}",
              file=sys.stderr)
        return 2

    if output_dir.exists() and not keep:
        # 仅清理目录内容，保留目录本身。
        for child in output_dir.iterdir():
            if child.is_dir() and not child.is_symlink():
                shutil.rmtree(child)
            else:
                child.unlink()

    output_dir.mkdir(parents=True, exist_ok=True)
    return 0


@timed("run_all_workbooks.main")
def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    repo = Path(__file__).resolve().parent
    code_dir = repo / 'Code'

    input_dir = _resolve_under_repo(repo, args.input_dir)
    if not input_dir.is_dir():
        print(f"[ERROR] input dir not found: {input_dir}", file=sys.stderr)
        return 2

    output_dir = _resolve_under_repo(repo, args.output_dir)
    if input_dir.resolve() == output_dir.resolve():
        print(f"[ERROR] input dir and output dir must differ: {input_dir}",
              file=sys.stderr)
        return 2

    rc = _prepare_output_dir(output_dir, repo, keep=args.keep_output)
    if rc != 0:
        return rc

    print(f"[INPUT]  {input_dir}")
    print(f"[OUTPUT] {output_dir}" + ("  (kept)" if args.keep_output else "  (cleaned)"))

    folder_to_wb: list[tuple[Path, str]] = []
    for folder in sorted(p for p in code_dir.iterdir() if p.is_dir()):
        wb = extract_workbook_from_folder(folder)
        if wb is None:
            continue
        folder_to_wb.append((folder, wb))

    folder_to_wb.sort(key=lambda x: workbook_sort_key(x[1]))

    child_env = os.environ.copy()
    child_env["GMF_INPUT_DIR"] = str(input_dir)
    child_env["GMF_OUTPUT_DIR"] = str(output_dir)

    sheet_runner = code_dir / '_sheet_runner.py'
    for idx, (folder, wb) in enumerate(folder_to_wb, 1):
        print(f"\n=== [{idx:02d}/{len(folder_to_wb):02d}] {wb} ===")
        result = subprocess.run(
            [sys.executable, str(sheet_runner), str(folder)],
            cwd=repo, env=child_env,
        )
        if result.returncode != 0:
            print(f"[STOP] {folder.name} exit={result.returncode}")
            return result.returncode

    print("\n[ALL DONE] all workbook runners completed")

    _postprocess_outputs(output_dir)
    _freeze_output_formulas(output_dir)

    print(f"[OUTPUT] results saved under: {output_dir}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
